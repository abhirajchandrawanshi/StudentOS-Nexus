import io
import csv
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_dsa_spreadsheet_bytes(questions: List[Dict[str, Any]], sheet_title: str = "DSA Preparation Sheet") -> tuple[bytes, str, str]:
    """
    Compiles problem metadata into spreadsheet bytes.
    If openpyxl is installed, generates a beautifully formatted Excel (.xlsx) file
    with columns, custom colors, auto-fit columns, and validation cells.
    Otherwise, generates a clean CSV (.csv) file.
    
    Returns a tuple of: (file_bytes, mime_type, file_extension)
    """
    
    # ─── 1. Format using openpyxl (Aesthetic Excel Spreadsheet) ─────────
    if OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DSA Prep"
        
        # Grid lines visible
        ws.views.sheetView[0].showGridLines = True
        
        # Styles definition
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid") # Dark indigo #1e1b4b
        
        cell_font = Font(name="Segoe UI", size=10)
        link_font = Font(name="Segoe UI", size=10, color="2563EB", underline="single") # Blue hyperlink
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Difficulty fills
        diff_fills = {
            "Easy": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),   # emerald green-100
            "Medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), # amber yellow-100
            "Hard": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")     # rose red-100
        }
        diff_fonts = {
            "Easy": Font(name="Segoe UI", size=10, color="065F46", bold=True),
            "Medium": Font(name="Segoe UI", size=10, color="92400E", bold=True),
            "Hard": Font(name="Segoe UI", size=10, color="991B1B", bold=True)
        }

        # Header Row
        headers = ["Q. ID", "Problem Title", "Topic Track", "Difficulty", "Company Relevance", "LeetCode URL", "My Mastery Status"]
        ws.append([]) # spacer row
        ws.append(["", sheet_title]) # title banner row
        
        # Format title banner
        ws.cell(row=2, column=2).font = Font(name="Segoe UI", size=15, bold=True, color="1E1B4B")
        
        # Append main headers
        ws.append(headers)
        header_row_idx = 4
        
        # Format headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = cell_border
        
        ws.row_dimensions[header_row_idx].height = 28
        
        # Add Data rows
        current_row = 5
        for idx, q in enumerate(questions):
            q_id = q.get("questionId") or q.get("id") or (idx + 1)
            title = q.get("title", "")
            topic = q.get("topic", "")
            diff = q.get("difficulty", "Medium")
            companies = ", ".join(q.get("companyTags", []))
            url = q.get("url", "")
            status = q.get("masteryStatus", "unsolved").upper()
            
            row_data = [q_id, title, topic, diff, companies, "Solve on LeetCode", status]
            ws.append(row_data)
            
            # Formatting Row cells
            ws.row_dimensions[current_row].height = 22
            
            # Q. ID
            cell = ws.cell(row=current_row, column=1)
            cell.font = cell_font
            cell.alignment = align_center
            cell.border = cell_border
            
            # Title
            cell = ws.cell(row=current_row, column=2)
            cell.font = cell_font
            cell.alignment = align_left
            cell.border = cell_border
            
            # Topic
            cell = ws.cell(row=current_row, column=3)
            cell.font = cell_font
            cell.alignment = align_center
            cell.border = cell_border
            
            # Difficulty
            cell = ws.cell(row=current_row, column=4)
            cell.font = diff_fonts.get(diff, cell_font)
            cell.fill = diff_fills.get(diff, PatternFill(fill_type=None))
            cell.alignment = align_center
            cell.border = cell_border
            
            # Companies
            cell = ws.cell(row=current_row, column=5)
            cell.font = cell_font
            cell.alignment = align_left
            cell.border = cell_border
            
            # URL Link
            cell = ws.cell(row=current_row, column=6)
            cell.value = "LeetCode Link"
            cell.hyperlink = url
            cell.font = link_font
            cell.alignment = align_center
            cell.border = cell_border
            
            # Status Dropdown tracker
            cell = ws.cell(row=current_row, column=7)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
            cell.alignment = align_center
            cell.border = cell_border
            
            current_row += 1

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip title banner in length calculations
                if cell.row == 2:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Specific overrides
        ws.column_dimensions["A"].width = 8   # ID
        ws.column_dimensions["B"].width = 30  # Title
        ws.column_dimensions["F"].width = 16  # Link
        ws.column_dimensions["G"].width = 18  # Status

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

    # ─── 2. Fallback to standard CSV Exporter ───────────────────────────
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(["Question ID", "Problem Title", "Topic Track", "Difficulty", "Target Companies", "LeetCode URL", "My Mastery Status"])
        
        # Rows
        for idx, q in enumerate(questions):
            q_id = q.get("questionId") or q.get("id") or (idx + 1)
            title = q.get("title", "")
            topic = q.get("topic", "")
            diff = q.get("difficulty", "Medium")
            companies = ", ".join(q.get("companyTags", []))
            url = q.get("url", "")
            status = q.get("masteryStatus", "unsolved")
            
            writer.writerow([q_id, title, topic, diff, companies, url, status])
            
        csv_data = output.getvalue().encode("utf-8")
        return csv_data, "text/csv", "csv"
